import pyautogui as pg
import datetime as dt
import openpyxl
from openpyxl import Workbook


# def create_excel():
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Rooms'
#     w2 = wb.create_sheet("IDs", 1)
#     for sheet in wb:
#         print(sheet.title)

#     dateCell = ws.cell(row=1, column=1, value='Date')
#     dateCell = ws.cell(row=1, column=2, value='Time')
#     dateCell = ws.cell(row=1, column=3, value='Room')
#     dateCell = ws.cell(row=1, column=4, value='Name')
#     roomCell = 2
#     nameCell = 2


#next step is to work on the excel writing and saving logic :)
#i think this needs to be done inside the respective methods :)

def textboxes():
    #captures locations of textboxes
    textb_list = list(pg.locateAllOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots\ox.png'))
    #locates and enters username
    pg.moveTo(textb_list[0])
    pg.click(textb_list[0])
    pg.typewrite("cosgrogc")
    #locates and enters password
    pg.moveTo(textb_list[1])
    pg.click(textb_list[1])
    pg.typewrite("yourpassword")  # replace with password

def signin():
    #locates and hits sign-in button
    sign_in_location = (pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots\sign_in2.png'))
    pg.moveTo(sign_in_location)
    pg.click(sign_in_location)

def reserveButton():
    #locate and click reserve a room
    pg._autoPause(.5,.5)
    reserve = pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots\serve.png')
    #pg.center(reserve)
    pg.moveTo(reserve)
    pg.click(reserve)

def arrows_weekday():
    pg._autoPause(.5, .5)
    arrow_list = list(pg.locateAllOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots\arrow.png'))
    pg.moveTo(arrow_list[0])
    pg.click(arrow_list[0])
    pg.press('enter')
    pg.moveTo(arrow_list[1])
    pg.click(arrow_list[1])
    pg.press('down'); pg.press('down')
    pg.press('enter')
    pg.moveTo(arrow_list[2])
    pg.click(arrow_list[2])
    start_at_1()
    pg.moveTo(arrow_list[3])
    pg.click(arrow_list[3])
    two_hours()
    click_search()

def arrows_weekend():
    pg._autoPause(.5, .5)
    arrow_list = list(pg.locateAllOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots\mk.png'))
    pg.moveTo(arrow_list[0])
    pg.click(arrow_list[0])
    pg.press('enter')
    pg.moveTo(arrow_list[1])
    pg.click(arrow_list[1])
    pg.press('down'); pg.press('down')
    pg.press('enter')
    pg.moveTo(arrow_list[2])
    pg.click(arrow_list[2])




def rooms():
    room_list = []
    if pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/22011.png'):
        room_list.append('2201')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2203.png'):
        room_list.append('2203')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2205.png'):
        room_list.append('2205')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2207.png'):
        room_list.append('2207')

    #elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2209.png'):
        #room_list.append('2209')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2211.png'):
        room_list.append('2211')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2213.png'):
        room_list.append('2213')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2215.png'):
        room_list.append('2215')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2203.png'):
        room_list.append('2217')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2219.png'):
        room_list.append('2219')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2221.png'):
        room_list.append('2221')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2223.png'):
        room_list.append('2223')

    elif pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/2225.png'):
        room_list.append('2225')

    #add error check here to do it in full window too possibly
    pg.moveTo(x =40, y =295, pause=1)
    pg.click()
    room = room_list[0]
    print(room)
    pg.moveTo(r'C:\Users\griff\projects\study-room-script\screenshots/select.png')
    pg.click()
    pg._autoPause(.5,.5)
    pg.moveTo(r'C:\Users\griff\projects\study-room-script\screenshots/confirm.png')
    pg.click()
    pg._autoPause(.5, .5)
    pg.moveTo(r'C:\Users\griff\projects\study-room-script\screenshots/return.png')
    pg.click()



def second_day():
    pg.press('down')
    pg.press('enter')

def start_at_1():
    x = 0
    while x < 10:
        pg.press('down')
        x += 1

    pg.press('enter')

def two_hours():
    x = 0
    while x < 7:
        pg.press('down')
        x += 1

    pg.press('enter')

def click_search():
    search = pg.locateOnScreen(r'C:\Users\griff\projects\study-room-script\screenshots/search.png')
    pg.moveTo(search)
    pg.click(search)


 
#textboxes()
#signin()
#reserveButton()
#arrows_weekday()  #if statement here to run weekend or friday or sat. sunday reg hours
#rooms()

#createExcel()   
